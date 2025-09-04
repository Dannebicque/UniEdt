# api_planning.py
# pip install fastapi uvicorn openpyxl
from fastapi import FastAPI, HTTPException, Path
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import Dict, Tuple, List, Optional
from io import BytesIO
from pathlib import Path
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

# -------- Paramètres fichiers (adapte les chemins à ton projet) --------
DATA_DIR = Path("../data-GEA")
CONFIG_PATH = DATA_DIR / "data_globale.json"          # semestres / groupes
COURSES_DIR = DATA_DIR / "cours"                      # contient cours_1.json, cours_2.json, ...

# -------- Paramètres par défaut --------
DEFAULT_DAYS = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]
DEFAULT_TIMES = ["8h00", "9h30", "11h00", "14h00", "15h30", "17h00"]

# Couleur de la ligne “Pause méridienne” (douce, lisible)
PAUSE_FILL = PatternFill("solid", fgColor="FFF7E7B6")  # AARRGGBB (jaune doux)
PAUSE_FONT = Font(bold=True, italic=True)

# -------- Utilitaires --------
thin = Side(border_style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def argb(color_hex: str) -> str:
    """#RRGGBB -> AARRGGBB (opaque)"""
    if not color_hex:
        return ""
    c = color_hex.replace("#", "").upper()
    if len(c) == 6:
        return "FF" + c
    if len(c) == 8:
        return c
    return ""

def ensure_borders(ws, row: int, col_start: int, col_end: int):
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = BORDER

def merge_or_single(ws, row: int, c1: int, c2: int):
    if c2 > c1:
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)

# -------- Génération (avec pause méridienne) --------
def build_sheet(config: Dict, courses: List[Dict], DAYS: List[str], TIMES: List[str]) -> Workbook:
    print(config)
    wb = Workbook()
    ws = wb.active
    ws.title = "Planning"

    # Col A pour les heures
    ws.column_dimensions["A"].width = 12
    ws.cell(row=1, column=1, value="Heure").font = Font(bold=True)

    # Colonnes par semestre (dans l'ordre du JSON)
    semester_col_ranges: Dict[str, Tuple[int, int]] = {}
    col = 2  # on commence en B
    for sem, info in config["semesters"].items():
        groupes_tp = list(info.get("groupesTp", {}).values())
        sem_color = argb(info.get("color", ""))

        start_col = col
        for g in groupes_tp:
            ws.cell(row=1, column=col, value=f"{sem}-{g}")
            ws.column_dimensions[get_column_letter(col)].width = 20
            head = ws.cell(row=1, column=col)
            if sem_color:
                head.fill = PatternFill("solid", fgColor=sem_color)
            head.font = Font(bold=True, color="FFFFFFFF" if sem_color else "000000")
            head.alignment = Alignment(horizontal="center", vertical="center")
            head.border = BORDER
            col += 1
        end_col = col - 1
        semester_col_ranges[sem] = (start_col, end_col)

    # Blocs jour + créneaux (avec pause entre créneau 3 et 4)
    row = 2
    for d in DAYS:
        # titre jour fusionné
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col - 1)
        day_cell = ws.cell(row=row, column=1, value=d)
        day_cell.font = Font(bold=True)
        day_cell.alignment = Alignment(horizontal="center", vertical="center")
        day_cell.fill = PatternFill("solid", fgColor="E6F2FF")
        ensure_borders(ws, row, 1, col - 1)
        row += 1

        # créneaux
        for idx, t in enumerate(TIMES, start=1):
            ctime = ws.cell(row=row, column=1, value=t)
            ctime.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 65
            ensure_borders(ws, row, 1, col - 1)
            row += 1

            # après le 3e créneau: insérer la pause méridienne
            if idx == 3:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col - 1)
                pause_cell = ws.cell(row=row, column=1, value="Pause méridienne")
                pause_cell.alignment = Alignment(horizontal="center", vertical="center")
                pause_cell.fill = PAUSE_FILL
                pause_cell.font = PAUSE_FONT
                ws.row_dimensions[row].height = 18
                ensure_borders(ws, row, 1, col - 1)
                row += 1

    # Helper pour retrouver la ligne (day, creneau)
    # IMPORTANT: on a ajouté 1 ligne “pause” après le créneau 3 => décalage pour 4,5,6
    def find_row(day: str, slot_index: int) -> int:
        r = 2
        while r <= ws.max_row:
            if ws.cell(r, 1).value == day:
                base = r + 1  # première ligne de créneau
                i = max(1, min(len(TIMES), int(slot_index)))
                offset = (i - 1) + (1 if i >= 4 else 0)  # +1 si on passe la pause
                return base + offset
            # sauter au bloc suivant : 1 (titre jour) + len(TIMES) + 1 (pause)
            r += 1 + len(TIMES) + 1
        return 0

    # Placer les cours
    for ev in courses:
        sem = ev.get("semester")
        if sem not in semester_col_ranges:
            continue

        day = ev.get("date")
        slot = int(ev.get("creneau") or 0)
        row_target = find_row(day, slot)
        if not row_target:
            continue

        start_col, end_col = semester_col_ranges[sem]
        typ = (ev.get("type") or "").upper().strip()

        # span: TP=1 case, TD=2 par défaut si non fourni, CM=toutes les cases
        span = int(ev.get("groupCount") or (2 if typ == "TD" else 1))
        gindex = int(ev.get("groupIndex") or 1)

        if typ == "CM":
            c1, c2 = start_col, end_col
        else:
            c1 = start_col + max(0, gindex - 1)
            c2 = min(end_col, c1 + span - 1)

        # Texte affiché
        parts = []
        if ev.get("heureDebut"): parts.append(f"Début : {ev['heureDebut']}")
        if ev.get("matiere"): parts.append(str(ev["matiere"]))
        if ev.get("professor"): parts.append(str(ev["professor"]))
        if ev.get("room"): parts.append(str(ev["room"]))
        duree = ev.get("duree")
        if duree:
            try:
                d = float(duree)
                h, m = int(d), int(round((d - int(d)) * 60))
                parts.append(f"Durée : {h}h{m:02d}" if m else f"Durée : {h}h")
            except Exception:
                pass
        value = "\n".join(parts)

        # Couleur (event > semestre)
        cell_color = argb(ev.get("color", "")) or argb(config["semesters"][sem].get("color", ""))

        merge_or_single(ws, row_target, c1, c2)
        cell = ws.cell(row=row_target, column=c1, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if cell_color:
            cell.fill = PatternFill("solid", fgColor=cell_color)
            cell.font = Font(color="FFFFFFFF")
        ensure_borders(ws, row_target, c1, c2)

    # Gèle sous l’en-tête
    ws.freeze_panes = ws["B3"]
    return wb

def generate_planning_bytes(config: Dict, courses: List[Dict], days=None, times=None) -> bytes:
    DAYS = days or DEFAULT_DAYS
    TIMES = times or DEFAULT_TIMES
    wb = build_sheet(config, courses, DAYS, TIMES)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


