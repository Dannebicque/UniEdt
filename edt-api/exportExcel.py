# planning_from_json.py
# pip install openpyxl

import json
from pathlib import Path
from typing import Dict, Tuple, List
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

# -------- Paramètres --------
DAYS = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]
# mapping 1..N -> libellé créneau
TIMES = ["8h00", "9h30", "11h00", "14h00", "15h30", "17h00"]

INPUT_CONFIG = Path("../data-GEA/data_globale.json")  # semestres / groupes
INPUT_COURSES = Path("../data-GEA/cours/cours_2.json")      # vos cours
OUTPUT_XLSX = Path("planning_semaine_2.xlsx")

# -------- Utilitaires --------
thin = Side(border_style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def time_from_index(idx: int) -> str:
    """1 -> 8h00, 2 -> 9h30, ... (borne dans la liste TIMES)"""
    if not idx:
        return ""
    i = max(1, min(len(TIMES), int(idx)))
    return TIMES[i - 1]

def argb(color_hex: str) -> str:
    """#RRGGBB -> AARRGGBB (pleinement opaque)."""
    if not color_hex:
        return ""
    c = color_hex.replace("#", "").upper()
    if len(c) == 6:
        return "FF" + c
    if len(c) == 8:
        return c
    return ""

def ensure_borders(ws, row: int, col_start: int, col_end: int):
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = BORDER

def merge_or_single(ws, row: int, c1: int, c2: int):
    if c2 > c1:
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)

# -------- Construction --------
def build_sheet(config: Dict, courses: List[Dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Planning"

    # Col A pour les heures
    ws.column_dimensions["A"].width = 12
    ws.cell(row=1, column=1, value="Heure").font = Font(bold=True)

    # Colonnes par semestre (dans l'ordre présent dans le JSON)
    semester_col_ranges: Dict[str, Tuple[int, int]] = {}
    col = 2  # on commence en B
    for sem, info in config["semesters"].items():
        groupes_tp = list(info.get("groupesTp", {}).values())
        sem_color = argb(info.get("color", ""))

        start_col = col
        for g in groupes_tp:
            ws.cell(row=1, column=col, value=f"{sem}-{g}")
            ws.column_dimensions[get_column_letter(col)].width = 20
            # teinte d'entête par semestre
            head = ws.cell(row=1, column=col)
            if sem_color:
                head.fill = PatternFill("solid", fgColor=sem_color)
            head.font = Font(bold=True, color="FFFFFFFF" if sem_color else "000000")
            head.alignment = Alignment(horizontal="center", vertical="center")
            head.border = BORDER
            col += 1
        end_col = col - 1
        semester_col_ranges[sem] = (start_col, end_col)

    # Construire les blocs Jours + Créneaux
    row = 2
    for d in DAYS:
        # ligne de titre jour fusionnée
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col - 1)
        day_cell = ws.cell(row=row, column=1, value=d)
        day_cell.font = Font(bold=True)
        day_cell.alignment = Alignment(horizontal="center", vertical="center")
        day_cell.fill = PatternFill("solid", fgColor="E6F2FF")
        ensure_borders(ws, row, 1, col - 1)
        row += 1

        # lignes de créneaux
        for t in TIMES:
            ctime = ws.cell(row=row, column=1, value=t)
            ctime.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 65  # Hauteur personnalisée (ex: 40 points)
            ensure_borders(ws, row, 1, col - 1)
            row += 1

    # Helper: retrouver la ligne (day, creneau)
    def find_row(day: str, slot_index: int) -> int:
        # on scanne les blocs jour
        r = 2
        while r <= ws.max_row:
            if ws.cell(r, 1).value == day:
                # la ligne suivante correspond à TIMES[0]
                base = r + 1
                i = max(1, min(len(TIMES), int(slot_index)))
                return base + (i - 1)
            # sauter au bloc suivant : 1 ligne (titre) + len(TIMES)
            r += 1 + len(TIMES)
        return 0

    # Placer les cours
    for ev in courses:
        sem = ev.get("semester")
        if sem not in semester_col_ranges:
            continue  # semestre non configuré

        day = ev.get("date")  # "Lundi", "Mardi", ...
        slot = int(ev.get("creneau") or 0)
        row_target = find_row(day, slot)
        if not row_target:
            continue

        start_col, end_col = semester_col_ranges[sem]
        typ = (ev.get("type") or "").upper().strip()
        # Par défaut (TP) 1 case
        span = int(ev.get("groupCount") or (2 if typ == "TD" else 1))
        gindex = int(ev.get("groupIndex") or 1)

        if typ == "CM":
            c1, c2 = start_col, end_col
        else:
            c1 = start_col + max(0, gindex - 1)
            c2 = min(end_col, c1 + span - 1)

        # Texte affiché
        lines = ""
        if ev.get("heureDebut"):
            lines = f"Début : {ev['heureDebut']}"

        # si lines non vide ajouter \n
        lines = lines + "\n" if lines else ""
        lines = lines + str(ev.get("matiere") or "")
        lines = lines + "\n" if lines else ""
        if ev["professor"]:
            lines = lines + (str(ev["professor"])) + "\n"
        if ev["room"]:
            lines = lines + (str(ev["room"])) + "\n"

        duree = ev.get("duree")
        if duree:
            heures = int(float(duree))
            minutes = int(round((float(duree) - heures) * 60))
            if minutes:
                lines = lines + "Durée : " + (f"{heures}h{minutes:02d}")
            else:
                lines = lines + "Durée : " + (f"{heures}h")

        value = lines

        # Couleur (priorité event, sinon couleur du semestre)
        cell_color = argb(ev.get("color", "")) or argb(config["semesters"][sem].get("color", ""))

        # Ecriture (on suppose pas de chevauchement)
        merge_or_single(ws, row_target, c1, c2)
        cell = ws.cell(row=row_target, column=c1, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if cell_color:
            cell.fill = PatternFill("solid", fgColor=cell_color)
            # si fond sombre on met police blanche
            cell.font = Font(color="FFFFFFFF")
        # borders pour la zone fusionnée
        ensure_borders(ws, row_target, c1, c2)

    # Gèle en haut à gauche sous l’entête
    ws.freeze_panes = ws["B3"]
    return wb

def main():
    with INPUT_CONFIG.open("r", encoding="utf-8") as f:
        config = json.load(f)

    with INPUT_COURSES.open("r", encoding="utf-8") as f:
        courses = json.load(f)

    wb = build_sheet(config, courses)
    wb.save(OUTPUT_XLSX)
    print(f"OK → {OUTPUT_XLSX.resolve()}")

if __name__ == "__main__":
    main()
